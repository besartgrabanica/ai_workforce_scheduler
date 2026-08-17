"""
Integration coverage for the /import-mappings route (app.py) — the
admin-gated confirm/edit UI from docs/specs/2026-08-import-mapping-detection.md.
Follows tests/test_dashboard_routes.py's pattern: call the view function
directly inside a test_request_context nested under the ctx fixture's app
context (g persists across the nesting — verified empirically), rather than
a full login/session HTTP round trip, which nothing else in this suite does
either. Combines `ctx` (project DB) with `identity_ctx` (identity DB) since
this route touches both (RBAC via a real IdentityUser/ProjectRole).

Every case that could reach Tier 3 monkeypatches
scheduler.ai_import_mapping.propose_mapping directly — this route imports it
with a local `from ... import` inside the view function, re-resolving on
every call, so patching the source module's attribute (not a copy bound at
app.py's own import time) is what actually takes effect here. Same real
.env ANTHROPIC_API_KEY concern as tests/test_import_mapping.py.
"""
import io

import openpyxl
import pytest
from flask import g

import scheduler.ai_import_mapping as aim
from app import app as flask_app
from app import db
from scheduler.models import IdentityUser, ImportMapping, ProjectRole


@pytest.fixture(autouse=True)
def _redirect_uploads_to_tmp(tmp_path, monkeypatch):
    """The route under test saves every uploaded file via app.upload_folder()
    (project-keyed, under the repo's real uploads/ dir) — redirect that to a
    throwaway tmp_path here so these tests don't litter uploads/eon/ with
    generated mapping_*.xlsx files on every run."""
    monkeypatch.setattr('app.upload_folder', lambda: str(tmp_path))


def _make_user(username, role, project='eon'):
    # project='eon' (not 'test', despite the ctx fixture's own default
    # g.active_project) because scheduler.parsers._REGISTRY only has a
    # module registered for 'eon' — get_parser_module('test') returns None,
    # which the route correctly (but unhelpfully, for a test) reports as
    # "this project's parser doesn't support that file type."
    user = IdentityUser(username=username, email=f'{username}@example.com')
    user.set_password('irrelevant')
    db.session.add(user)
    db.session.commit()
    db.session.add(ProjectRole(user_id=user.id, project=project, role=role))
    db.session.commit()
    return user


def _tfc_bytes_without_gesamt():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KiKxxl'
    ws.cell(row=9, column=1, value='Juni')
    ws.cell(row=9, column=2, value=23)
    ws.cell(row=9, column=3, value='Mo')
    ws.cell(row=9, column=4, value='2026-06-01')
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _tfc_bytes_reformatted_with_real_data_at(col0):
    """Reformatted (no 'Gesamt' label anywhere, so Tier 1 still warns) but
    with real non-zero volume at col0 — the shape a correct manual mapping
    of this exact file looks like, as opposed to _tfc_bytes_without_gesamt
    (nothing at any column, which any mapping now correctly fails to
    validate against, per eon.py's "produced no volume" check)."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KiKxxl'
    ws.cell(row=9, column=1, value='Juni')
    ws.cell(row=9, column=2, value=23)
    ws.cell(row=9, column=3, value='Mo')
    ws.cell(row=9, column=4, value='2026-06-01')
    ws.cell(row=9, column=col0 + 1, value=55.0)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _tfc_bytes_with_gesamt(col0):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KiKxxl'
    ws.cell(row=7, column=col0 + 1, value='Gesamt')
    ws.cell(row=9, column=1, value='Juni')
    ws.cell(row=9, column=2, value=23)
    ws.cell(row=9, column=3, value='Mo')
    from datetime import date
    ws.cell(row=9, column=4, value=date(2026, 6, 1))
    ws.cell(row=9, column=col0 + 1, value=120.5)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _post(data, user, project='eon'):
    with flask_app.test_request_context(
        '/import-mappings', method='POST', data=data, content_type='multipart/form-data',
    ):
        g.current_user = user
        g.active_project = project
        return flask_app.view_functions['import_mappings']()


def _get(user, project='eon'):
    with flask_app.test_request_context('/import-mappings'):
        g.current_user = user
        g.active_project = project
        return flask_app.view_functions['import_mappings']()


# ── RBAC ─────────────────────────────────────────────────────────────────────

def test_get_blocked_for_non_admin(ctx, identity_ctx):
    viewer = _make_user('viewer1', 'viewer')

    resp = _get(viewer)

    assert resp.status_code == 302
    assert ImportMapping.query.count() == 0


def test_post_blocked_for_non_admin_even_with_a_valid_mapping(ctx, identity_ctx, monkeypatch):
    editor = _make_user('editor1', 'editor')
    data = {
        'file_type': 'tfc_forecast', 'gesamt_col': '12',
        'mapping_file': (_tfc_bytes_reformatted_with_real_data_at(12), 'tfc.xlsx'),
    }

    resp = _post(data, editor)

    assert resp.status_code == 302
    assert ImportMapping.query.count() == 0


def test_get_allowed_for_admin(ctx, identity_ctx):
    # A rendered template comes back as a plain string when the view is
    # called directly (bypassing Flask's own dispatch, which normally wraps
    # it into a Response) — only redirect() below returns a real Response.
    admin = _make_user('admin1', 'admin')

    resp = _get(admin)

    assert isinstance(resp, str)
    assert 'Import Mappings' in resp


# ── manual entry (zero AI calls) ─────────────────────────────────────────────

def test_manual_entry_saves_mapping_without_calling_ai(ctx, identity_ctx, monkeypatch):
    admin = _make_user('admin2', 'admin')
    calls = []
    monkeypatch.setattr(aim, 'propose_mapping', lambda *a, **k: calls.append(1) or None)

    data = {
        'file_type': 'tfc_forecast', 'gesamt_col': '12',
        'mapping_file': (_tfc_bytes_reformatted_with_real_data_at(12), 'tfc.xlsx'),
    }
    resp = _post(data, admin)

    assert resp.status_code == 302
    assert calls == []  # propose_mapping never invoked on the manual-entry path
    row = ImportMapping.query.one()
    assert row.file_type == 'tfc_forecast'
    assert row.resolution_source == 'manual'
    assert row.confirmed_by == 'admin2'


def test_manual_entry_with_bad_mapping_saves_nothing(ctx, identity_ctx):
    admin = _make_user('admin3', 'admin')
    data = {
        # Points at a genuinely empty column — zero volume on every day
        # triggers eon.py's "produced no volume" check (see save_confirmed_mapping).
        'file_type': 'tfc_forecast', 'gesamt_col': '999',
        'mapping_file': (_tfc_bytes_without_gesamt(), 'tfc.xlsx'),
    }

    resp = _post(data, admin)

    assert isinstance(resp, str)  # re-rendered with an error, not redirected
    assert ImportMapping.query.count() == 0


# ── already-clean file: no mapping needed, no AI call ────────────────────────

def test_clean_file_with_blank_fields_needs_no_mapping_and_calls_no_ai(ctx, identity_ctx, monkeypatch):
    admin = _make_user('admin4', 'admin')
    calls = []
    monkeypatch.setattr(aim, 'propose_mapping', lambda *a, **k: calls.append(1) or None)

    data = {
        'file_type': 'tfc_forecast', 'gesamt_col': '',
        'mapping_file': (_tfc_bytes_with_gesamt(10), 'tfc.xlsx'),
    }
    resp = _post(data, admin)

    assert isinstance(resp, str)
    assert calls == []  # Tier 1 already resolves this file — Tier 3 must never fire
    assert ImportMapping.query.count() == 0


# ── AI suggestion path ───────────────────────────────────────────────────────

def test_blank_fields_on_a_reformatted_file_returns_ai_suggestion(ctx, identity_ctx, monkeypatch):
    admin = _make_user('admin5', 'admin')
    fake_proposal = {'mapping': {'gesamt_col': 12}, 'confidence': 0.8, 'rationale': 'moved'}
    monkeypatch.setattr(aim, 'propose_mapping', lambda file_type, snapshot: fake_proposal)

    data = {
        'file_type': 'tfc_forecast', 'gesamt_col': '',
        'mapping_file': (_tfc_bytes_without_gesamt(), 'tfc.xlsx'),
    }
    resp = _post(data, admin)

    assert isinstance(resp, str)
    assert 'moved' in resp  # the stubbed rationale shows up in the review page
    assert ImportMapping.query.count() == 0  # a suggestion is never auto-saved


def test_confirming_an_ai_suggestion_records_ai_confirmed_provenance(ctx, identity_ctx):
    admin = _make_user('admin6', 'admin')
    # Simulates the second submit of the two-step flow: the admin re-uploaded
    # the file and kept the AI-suggested value, so ai_assisted/confidence/
    # rationale ride along as hidden fields exactly as the template renders them.
    data = {
        'file_type': 'tfc_forecast', 'gesamt_col': '12',
        'ai_assisted': '1', 'ai_confidence': '0.8', 'ai_rationale': 'moved right',
        'mapping_file': (_tfc_bytes_reformatted_with_real_data_at(12), 'tfc.xlsx'),
    }
    resp = _post(data, admin)

    assert resp.status_code == 302
    row = ImportMapping.query.one()
    assert row.resolution_source == 'ai_confirmed'
    assert row.confidence == pytest.approx(0.8)
    assert row.rationale == 'moved right'
    assert row.confirmed_by == 'admin6'
