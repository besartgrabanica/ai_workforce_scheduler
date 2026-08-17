"""
Regression suite for scheduler/parsers/eon.py. No real client files exist in
this repo (they're confidential source data, never committed) — each test
builds a minimal synthetic .xlsx workbook in-memory/on-disk that matches the
exact column-offset layout documented in eon.py's own docstrings and module-
level constants, then exercises the parser against it. Covers both the
happy path and the "expected structure not found" warning path for each of
the four parser functions, since ImportLog/the warning-flash UI depends on
those warnings actually firing rather than silently guessing or crashing.
"""
from datetime import date

import openpyxl
import pytest

from scheduler.parsers.eon import (
    parse_abnahme_de_file,
    parse_employee_file,
    parse_forecast_calculation_file,
    parse_tfc_file,
)


def _set(ws, row, col0, value):
    """col0 is the 0-based column index eon.py's row-tuple indexing uses."""
    ws.cell(row=row, column=col0 + 1, value=value)


def _save(wb, tmp_path, name='test.xlsx'):
    path = str(tmp_path / name)
    wb.save(path)
    return path


# ── parse_tfc_file ───────────────────────────────────────────────────────────

def test_tfc_parses_totals_and_slots_when_gesamt_header_found(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KiKxxl'
    gesamt_col = 10  # 0-based, deliberately clear of the reserved 0-3 columns

    _set(ws, 7, gesamt_col, 'Gesamt')

    _set(ws, 9, 0, 'Juni')                  # Monat
    _set(ws, 9, 1, 23)                      # KW
    _set(ws, 9, 2, 'Mo')                    # day tag
    _set(ws, 9, 3, date(2026, 6, 1))        # date
    _set(ws, 9, gesamt_col + 0, 120.5)      # sync total
    _set(ws, 9, gesamt_col + 1, 5.0)        # sync slot 08:00
    _set(ws, 9, gesamt_col + 25, 45.0)      # async total
    _set(ws, 9, gesamt_col + 50, 10.0)      # chat total

    # A row with no Monat must be skipped entirely, even with a valid date.
    _set(ws, 10, 3, date(2026, 6, 2))

    results, warnings = parse_tfc_file(_save(wb, tmp_path))

    assert warnings == []
    assert len(results) == 1
    day = results[0]
    assert day['date'] == date(2026, 6, 1)
    assert day['day_tag'] == 'Mo'
    assert day['kw'] == 23
    assert day['total_sync'] == 120.5
    assert day['total_async'] == 45.0
    assert day['total_chat'] == 10.0
    assert day['slots']['08:00']['sync'] == 5.0
    assert day['slots']['08:00']['async'] == 0.0
    assert day['slots']['08:30']['sync'] == 0.0


def test_tfc_falls_back_and_warns_when_gesamt_header_missing(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KiKxxl'
    # Row 7 deliberately has no 'Gesamt' anywhere.
    _set(ws, 9, 0, 'Juni')
    _set(ws, 9, 1, 23)
    _set(ws, 9, 2, 'Mo')
    _set(ws, 9, 3, date(2026, 6, 1))

    results, warnings = parse_tfc_file(_save(wb, tmp_path))

    assert len(results) == 1  # a row was still readable, just with 0-valued totals
    assert results[0]['total_sync'] == 0.0
    assert any('Gesamt' in w for w in warnings)
    assert not any('No daily rows' in w for w in warnings)


def test_tfc_raises_when_sheet_missing(tmp_path):
    wb = openpyxl.Workbook()  # default sheet named 'Sheet', not 'KiKxxl'
    with pytest.raises(ValueError, match='KiKxxl'):
        parse_tfc_file(_save(wb, tmp_path))


def test_tfc_warns_when_given_mapping_column_has_no_volume(tmp_path):
    """A wrong-but-plausible gesamt_col (still finds date rows, just at an
    empty column) must not silently validate as clean — this is what
    scheduler.import_mapping.save_confirmed_mapping relies on to refuse a
    fat-fingered manual entry. Only checked when mapping is given; Tier 1
    (mapping=None) never hits this, so it can't affect the unmodified-file
    behavior covered by the tests above."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KiKxxl'
    _set(ws, 9, 0, 'Juni')
    _set(ws, 9, 1, 23)
    _set(ws, 9, 2, 'Mo')
    _set(ws, 9, 3, date(2026, 6, 1))
    # Deliberately nothing at column 50 — the mapping under test.

    results, warnings = parse_tfc_file(_save(wb, tmp_path), mapping={'gesamt_col': 50})

    assert len(results) == 1
    assert any('produced no volume' in w for w in warnings)


# ── parse_abnahme_de_file ────────────────────────────────────────────────────

def test_abnahme_de_reads_last_column_after_datum_header(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _set(ws, 1, 1, 'Datum')                    # header row, col index 1
    _set(ws, 1, 4, 'Abnahme DE netto')          # rightmost non-None -> value column
    _set(ws, 2, 1, date(2026, 6, 1))
    _set(ws, 2, 4, 250.75)

    result, warnings = parse_abnahme_de_file(_save(wb, tmp_path))

    assert warnings == []
    assert result == {date(2026, 6, 1): 250.75}


def test_abnahme_de_warns_when_datum_header_missing(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _set(ws, 1, 1, 'Not Datum')
    _set(ws, 2, 1, date(2026, 6, 1))
    _set(ws, 2, 4, 250.75)

    result, warnings = parse_abnahme_de_file(_save(wb, tmp_path))

    assert result == {}
    assert any('Datum' in w for w in warnings)


def test_abnahme_de_warns_when_given_mapping_columns_produce_no_data(tmp_path):
    """Mirrors the tfc_forecast 'no volume' check: a wrong-but-plausible
    date_col/value_col (no exception, just nothing there) must not silently
    validate as clean. Only checked when mapping is given."""
    wb = openpyxl.Workbook()
    ws = wb.active
    _set(ws, 1, 0, 'unrelated header')
    _set(ws, 2, 1, date(2026, 6, 1))
    _set(ws, 2, 4, 250.75)
    # date_col/value_col point at columns with nothing in them.

    result, warnings = parse_abnahme_de_file(_save(wb, tmp_path), mapping={'date_col': 9, 'value_col': 10})

    assert result == {}
    assert any('produced no data' in w for w in warnings)


# ── parse_forecast_calculation_file ─────────────────────────────────────────

def test_forecast_calculation_reads_kikxxl_column(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _set(ws, 8, 1, 'Datum')
    _set(ws, 8, 5, 'KiKxxl\nPrishtina')
    _set(ws, 10, 1, date(2026, 6, 1))
    _set(ws, 10, 5, 45.9)  # truncates toward zero via int(float(...))

    result, warnings = parse_forecast_calculation_file(_save(wb, tmp_path))

    assert warnings == []
    assert result == {date(2026, 6, 1): 45}


def test_forecast_calculation_warns_when_kikxxl_column_missing(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _set(ws, 8, 1, 'Datum')
    _set(ws, 10, 1, date(2026, 6, 1))

    result, warnings = parse_forecast_calculation_file(_save(wb, tmp_path))

    assert result == {}
    assert any('KiKxxl' in w for w in warnings)


# ── parse_employee_file ─────────────────────────────────────────────────────

def test_employee_file_parses_teams_and_weekend_holiday_flags(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _set(ws, 1, 0, 'Team A')
    _set(ws, 2, 1, 'Emri & Mbiemri')  # header row — must not become an employee

    _set(ws, 3, 1, 'Agron Krasniqi')
    _set(ws, 3, 2, 'some note')
    _set(ws, 3, 3, 'Yes')
    _set(ws, 3, 4, 'Yes')
    _set(ws, 3, 5, 0.8)

    _set(ws, 4, 1, 'Blerta Hoxha')
    _set(ws, 4, 3, 'Saturday only')
    _set(ws, 4, 4, 'No')
    _set(ws, 4, 5, 1.0)

    _set(ws, 5, 0, 'Team B')
    _set(ws, 6, 1, 'Dardan Berisha')
    _set(ws, 6, 3, 'Sunday only')

    _set(ws, 7, 1, 'Note:')  # must be skipped, not treated as an employee

    employees, warnings = parse_employee_file(_save(wb, tmp_path))

    assert warnings == []
    assert [e['name'] for e in employees] == ['Agron Krasniqi', 'Blerta Hoxha', 'Dardan Berisha']

    agron = employees[0]
    assert agron['team'] == 'Team A'
    assert agron['works_saturday'] == 'yes'
    assert agron['works_sunday'] == 'no'
    assert agron['works_holidays'] is True
    assert agron['fte_fraction'] == 0.8

    blerta = employees[1]
    assert blerta['works_saturday'] == 'yes'
    assert blerta['works_sunday'] == 'no'
    assert blerta['works_holidays'] is False

    dardan = employees[2]
    assert dardan['team'] == 'Team B'
    assert dardan['works_saturday'] == 'no'
    assert dardan['works_sunday'] == 'yes'
    assert dardan['fte_fraction'] is None


def test_employee_file_warns_when_header_row_never_seen(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    # A real-looking, full-width employee row, but the header text itself
    # never appears anywhere in the sheet (e.g. it was reworded upstream).
    _set(ws, 1, 0, 'Team A')
    _set(ws, 2, 1, 'Agron Krasniqi')
    _set(ws, 2, 2, '')
    _set(ws, 2, 3, 'Yes')
    _set(ws, 2, 4, 'No')
    _set(ws, 2, 5, 1.0)

    employees, warnings = parse_employee_file(_save(wb, tmp_path))

    assert len(employees) == 1
    assert any('Emri & Mbiemri' in w for w in warnings)


def test_employee_file_warns_when_no_employees_found(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _set(ws, 1, 1, 'Emri & Mbiemri')  # header seen, but nobody follows it
    _set(ws, 1, 5, None)  # keep the sheet the full 6-column width regardless

    employees, warnings = parse_employee_file(_save(wb, tmp_path))

    assert employees == []
    assert any('No employees' in w for w in warnings)
    assert not any('Emri & Mbiemri' in w for w in warnings)
