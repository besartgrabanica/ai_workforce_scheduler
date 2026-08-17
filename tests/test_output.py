"""
Regression suite for scheduler/output.py::export_schedule() — builds a real
Schedule + ShiftAssignment rows against the existing test fixtures and
inspects the returned workbook's actual cells, rather than mocking the
export. Also documents (via test_coverage_falls_back_to_hardcoded_default_
without_daily_forecast) a real inconsistency worth knowing about: this
module's no-forecast fallback (220/55/0, weekday/Saturday/Sunday) doesn't
match scheduler/algorithm.py's fallback, which reads the BusinessParam
defaults instead — see the plan notes for A1. Not fixed here, just pinned
down so it's visible instead of silently assumed.
"""
from datetime import date

import openpyxl

from app import db
from scheduler.models import DailyForecast, ForecastPeriod, Schedule, ShiftAssignment, ShiftTemplate
from scheduler.output import export_schedule
from tests.conftest import make_employee


def _load(buf):
    return openpyxl.load_workbook(buf)


def test_export_includes_employee_and_assignment_in_schedule_sheet(default_template):
    emp = make_employee('Agron Krasniqi', fte_percent=100, team='Support')
    sched = Schedule(name='Test', year=2026, month=6)
    db.session.add(sched)
    db.session.commit()

    tpl = ShiftTemplate.query.filter_by(is_default=True).first()
    db.session.add(ShiftAssignment(schedule_id=sched.id, employee_id=emp.id, date=date(2026, 6, 1),
                                    status='work', shift_start='08:00', shift_end='16:30',
                                    hours_worked=7.5, shift_template_id=tpl.id))
    db.session.add(ShiftAssignment(schedule_id=sched.id, employee_id=emp.id, date=date(2026, 6, 2),
                                    status='day_off'))
    db.session.commit()

    wb = _load(export_schedule(sched.id, project_display_name='Test Co'))

    assert wb.sheetnames == ['Schedule', 'Coverage']
    ws = wb['Schedule']
    assert ws['A1'].value == 'Test Co Workforce Schedule — June 2026'
    # Row 2 = header, row 3 = team separator ('Support'), row 4 = Agron's data row.
    assert ws.cell(row=2, column=1).value == 'Team'
    assert ws.cell(row=3, column=1).value == 'Support'
    assert ws.cell(row=4, column=2).value == 'Agron Krasniqi'
    assert ws.cell(row=4, column=3).value == '100%'
    # Day 1 (June 1, 2026 is a Monday) -> the work assignment's display code.
    assert ws.cell(row=4, column=4).value == '08:00–16:30'
    # Day 2 -> day_off's display code.
    assert ws.cell(row=4, column=5).value == 'OFF'


def test_coverage_uses_daily_forecast_when_present(default_template):
    emp = make_employee('Blerta Hoxha', fte_percent=100)
    sched = Schedule(name='Test', year=2026, month=6)
    db.session.add(sched)
    db.session.commit()

    tpl = ShiftTemplate.query.filter_by(is_default=True).first()
    db.session.add(ShiftAssignment(schedule_id=sched.id, employee_id=emp.id, date=date(2026, 6, 1),
                                    status='work', shift_start='08:00', shift_end='16:30',
                                    hours_worked=7.5, shift_template_id=tpl.id))

    period = ForecastPeriod(name='June', start_date=date(2026, 6, 1), end_date=date(2026, 6, 30))
    db.session.add(period)
    db.session.flush()
    db.session.add(DailyForecast(period_id=period.id, date=date(2026, 6, 1), required_ks_agents=3))
    db.session.commit()

    wb = _load(export_schedule(sched.id))
    ws = wb['Coverage']
    # Row 2 = June 1.
    assert ws.cell(row=2, column=1).value == '01.06.2026'
    assert ws.cell(row=2, column=3).value == 3        # required, from DailyForecast
    assert ws.cell(row=2, column=4).value == 1        # scheduled
    assert ws.cell(row=2, column=6).value == 1 - 3     # gap


def test_coverage_falls_back_to_hardcoded_default_without_daily_forecast(default_template):
    """No DailyForecast row exists for this schedule's month at all — the
    Coverage sheet falls back to output.py's own hardcoded defaults
    (220 weekday / 55 Saturday / 0 Sunday), independent of any BusinessParam
    the project may have configured. This is the exact fallback
    scheduler/algorithm.py does NOT use (it reads BusinessParam defaults
    instead) — pinned down here as documentation of the current behavior,
    not an endorsement of it."""
    sched = Schedule(name='Test', year=2026, month=6)
    db.session.add(sched)
    db.session.commit()

    wb = _load(export_schedule(sched.id))
    ws = wb['Coverage']

    # June 1, 2026 is a Monday (weekday).
    assert ws.cell(row=2, column=1).value == '01.06.2026'
    assert ws.cell(row=2, column=3).value == 220
    # June 6, 2026 is a Saturday -> row 7 (June 1 = row 2, so June 6 = row 7).
    assert ws.cell(row=7, column=3).value == 55
    # June 7, 2026 is a Sunday -> row 8.
    assert ws.cell(row=8, column=3).value == 0
