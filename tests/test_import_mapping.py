"""
Regression suite for scheduler/import_mapping.py — Tiers 1/2/3 of the
3-tier import resolution (docs/specs/2026-08-import-mapping-detection.md).
Builds minimal synthetic .xlsx workbooks the same way
tests/test_parsers_eon.py does, then exercises fingerprinting and the
deterministic/cached/ai-proposed resolution flow against them.

Every test that could otherwise fall through to Tier 3 monkeypatches
scheduler.import_mapping.propose_mapping — this repo's real .env carries a
live ANTHROPIC_API_KEY (loaded by app.py's load_dotenv() at import time),
so leaving Tier 3 unstubbed here would fire a real, billed Anthropic API
call on every test run. See tests/test_ai_import_mapping.py for
propose_mapping's own (also fully stubbed) coverage.
"""
import json
from datetime import date

import openpyxl

from app import db
from scheduler.import_mapping import (
    InvalidMapping,
    build_layout_snapshot,
    compute_layout_fingerprint,
    resolve_cached_mapping,
    resolve_import,
    save_confirmed_mapping,
)
from scheduler.models import ImportMapping
from scheduler.parsers.eon import parse_abnahme_de_file, parse_tfc_file


def _set(ws, row, col0, value):
    ws.cell(row=row, column=col0 + 1, value=value)


def _save(wb, tmp_path, name='test.xlsx'):
    path = str(tmp_path / name)
    wb.save(path)
    return path


def _tfc_workbook_with_gesamt(gesamt_col):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KiKxxl'
    _set(ws, 7, gesamt_col, 'Gesamt')
    _set(ws, 9, 0, 'Juni')
    _set(ws, 9, 1, 23)
    _set(ws, 9, 2, 'Mo')
    _set(ws, 9, 3, date(2026, 6, 1))
    _set(ws, 9, gesamt_col, 120.5)
    return wb


def _tfc_workbook_without_gesamt():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'KiKxxl'
    _set(ws, 9, 0, 'Juni')
    _set(ws, 9, 1, 23)
    _set(ws, 9, 2, 'Mo')
    _set(ws, 9, 3, date(2026, 6, 1))
    return wb


# ── compute_layout_fingerprint ──────────────────────────────────────────────

def test_fingerprint_is_stable_across_data_row_changes(tmp_path):
    wb1 = _tfc_workbook_with_gesamt(10)
    fp1 = compute_layout_fingerprint(_save(wb1, tmp_path, 'a.xlsx'))

    wb2 = _tfc_workbook_with_gesamt(10)
    _set(wb2.active, 9, 10, 999.0)  # different data value, same header layout
    fp2 = compute_layout_fingerprint(_save(wb2, tmp_path, 'b.xlsx'))

    assert fp1 == fp2


def test_fingerprint_changes_when_header_layout_changes(tmp_path):
    fp_at_10 = compute_layout_fingerprint(_save(_tfc_workbook_with_gesamt(10), tmp_path, 'a.xlsx'))
    fp_no_header = compute_layout_fingerprint(_save(_tfc_workbook_without_gesamt(), tmp_path, 'b.xlsx'))

    assert fp_at_10 != fp_no_header


# ── resolve_cached_mapping ───────────────────────────────────────────────────

def test_resolve_cached_mapping_misses_with_no_rows(ctx, tmp_path):
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)
    assert resolve_cached_mapping('tfc_forecast', path) is None


def test_resolve_cached_mapping_hits_on_matching_file_type_and_fingerprint(ctx, tmp_path):
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)
    fingerprint = compute_layout_fingerprint(path)
    db.session.add(ImportMapping(
        file_type='tfc_forecast', layout_fingerprint=fingerprint,
        mapping_data=json.dumps({'gesamt_col': 10}),
    ))
    db.session.commit()

    found = resolve_cached_mapping('tfc_forecast', path)
    assert found is not None
    assert json.loads(found.mapping_data) == {'gesamt_col': 10}


def test_resolve_cached_mapping_ignores_wrong_file_type(ctx, tmp_path):
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)
    fingerprint = compute_layout_fingerprint(path)
    db.session.add(ImportMapping(
        file_type='abnahme_de', layout_fingerprint=fingerprint,
        mapping_data=json.dumps({'date_col': 1, 'value_col': 4}),
    ))
    db.session.commit()

    assert resolve_cached_mapping('tfc_forecast', path) is None


# ── resolve_import ───────────────────────────────────────────────────────────

def test_resolve_import_uses_tier1_unchanged_when_clean(ctx, tmp_path):
    path = _save(_tfc_workbook_with_gesamt(10), tmp_path)

    result, warnings, tier, proposal = resolve_import('tfc_forecast', path, parse_tfc_file)

    assert tier == 'deterministic'
    assert warnings == []
    assert proposal is None
    assert result[0]['total_sync'] == 120.5
    assert ImportMapping.query.count() == 0  # Tier 2 never consulted on a clean Tier 1 hit


def test_resolve_import_falls_back_to_tier1_when_no_cache_and_no_ai_proposal(ctx, tmp_path, monkeypatch):
    monkeypatch.setattr('scheduler.import_mapping.propose_mapping', lambda file_type, snapshot: None)
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)

    result, warnings, tier, proposal = resolve_import('tfc_forecast', path, parse_tfc_file)

    assert tier == 'deterministic'
    assert proposal is None
    assert any('Gesamt' in w for w in warnings)


def test_resolve_import_applies_cached_mapping_on_tier1_miss(ctx, tmp_path):
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)

    # Reformatted file: header search fails, but the real Gesamt data lives at col 20.
    wb = openpyxl.load_workbook(path)
    ws = wb['KiKxxl']
    _set(ws, 9, 20, 77.0)
    wb.save(path)

    fingerprint = compute_layout_fingerprint(path)
    db.session.add(ImportMapping(
        file_type='tfc_forecast', layout_fingerprint=fingerprint,
        mapping_data=json.dumps({'gesamt_col': 20}),
    ))
    db.session.commit()

    # No propose_mapping stub needed: a clean Tier 2 hit returns before Tier 3
    # is ever consulted (asserted below), so this can't hit the real API.
    result, warnings, tier, proposal = resolve_import('tfc_forecast', path, parse_tfc_file)

    assert tier == 'cached'
    assert warnings == []
    assert proposal is None
    assert result[0]['total_sync'] == 77.0


def test_resolve_import_distrusts_cached_mapping_that_still_warns(ctx, tmp_path, monkeypatch):
    monkeypatch.setattr('scheduler.import_mapping.propose_mapping', lambda file_type, snapshot: None)
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)
    fingerprint = compute_layout_fingerprint(path)
    # A stale/bad cached mapping — 'gesamt_col' key deliberately absent, so
    # parse_tfc_file falls straight back to its own header search, which
    # still misses on this file and re-warns — falls through to Tier 3,
    # which (stubbed above) also has nothing to offer.
    db.session.add(ImportMapping(
        file_type='tfc_forecast', layout_fingerprint=fingerprint,
        mapping_data=json.dumps({}),
    ))
    db.session.commit()

    result, warnings, tier, proposal = resolve_import('tfc_forecast', path, parse_tfc_file)

    assert tier == 'deterministic'
    assert proposal is None
    assert any('Gesamt' in w for w in warnings)


def test_resolve_import_returns_ai_proposed_tier_without_applying_it(ctx, tmp_path, monkeypatch):
    """Tier 3 firing must never change what resolve_import hands back as
    parsed data — acceptance criterion 2 ("nothing downstream executes
    pre-confirm") is enforced by the caller trusting only
    result/warnings, which here must still be Tier 1's own degraded
    output, not anything derived from the (stubbed) proposal."""
    fake_proposal = {'mapping': {'gesamt_col': 99}, 'confidence': 0.75, 'rationale': 'looks right'}
    monkeypatch.setattr('scheduler.import_mapping.propose_mapping', lambda file_type, snapshot: fake_proposal)
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)

    result, warnings, tier, proposal = resolve_import('tfc_forecast', path, parse_tfc_file)

    assert tier == 'ai_proposed'
    assert proposal == fake_proposal
    assert any('Gesamt' in w for w in warnings)  # Tier 1's original warning, untouched
    assert result[0]['total_sync'] == 0.0  # Tier 1's own fallback value, NOT derived from gesamt_col=99


# ── build_layout_snapshot ────────────────────────────────────────────────────

def test_build_layout_snapshot_captures_nonempty_cells_and_stringifies_dates(tmp_path):
    wb = _tfc_workbook_with_gesamt(10)
    _set(wb.active, 20, 0, 'row 20 — beyond default max_rows=15, must be excluded')
    path = _save(wb, tmp_path)

    snapshot = build_layout_snapshot(path, 'tfc_forecast')

    assert {'row': 6, 'col': 10, 'value': 'Gesamt'} in snapshot
    # openpyxl round-trips a pure date() as a datetime with a zero time component.
    assert {'row': 8, 'col': 3, 'value': '2026-06-01T00:00:00'} in snapshot
    assert not any(c['row'] > 14 for c in snapshot)  # respects max_rows=15 default


def test_build_layout_snapshot_uses_active_sheet_for_abnahme_de(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _set(ws, 1, 0, 'unrelated header')
    _set(ws, 2, 3, 250.75)
    path = _save(wb, tmp_path)

    snapshot = build_layout_snapshot(path, 'abnahme_de')

    assert {'row': 0, 'col': 0, 'value': 'unrelated header'} in snapshot
    assert {'row': 1, 'col': 3, 'value': 250.75} in snapshot


# ── parser mapping overrides ─────────────────────────────────────────────────

def test_parse_tfc_file_with_mapping_skips_header_search(tmp_path):
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)
    wb = openpyxl.load_workbook(path)
    _set(wb['KiKxxl'], 9, 15, 42.0)
    wb.save(path)

    results, warnings = parse_tfc_file(path, mapping={'gesamt_col': 15})

    assert warnings == []
    assert results[0]['total_sync'] == 42.0


def test_parse_abnahme_de_file_with_mapping_skips_header_search(tmp_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    _set(ws, 1, 0, 'unrelated header')  # deliberately no 'Datum' anywhere
    _set(ws, 2, 0, date(2026, 6, 1))
    _set(ws, 2, 3, 300.0)
    path = _save(wb, tmp_path)

    result, warnings = parse_abnahme_de_file(path, mapping={'date_col': 0, 'value_col': 3})

    assert warnings == []
    assert result == {date(2026, 6, 1): 300.0}


# ── save_confirmed_mapping ───────────────────────────────────────────────────

def test_save_confirmed_mapping_creates_a_new_row(ctx, tmp_path):
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)
    wb = openpyxl.load_workbook(path)
    _set(wb['KiKxxl'], 9, 12, 88.0)
    wb.save(path)

    row = save_confirmed_mapping(
        file_type='tfc_forecast', filepath=path, mapping={'gesamt_col': 12},
        parse_fn=parse_tfc_file, resolution_source='manual', confirmed_by='besart',
    )

    assert row.id is not None
    assert row.file_type == 'tfc_forecast'
    assert row.layout_fingerprint == compute_layout_fingerprint(path)
    assert json.loads(row.mapping_data) == {'gesamt_col': 12}
    assert row.resolution_source == 'manual'
    assert row.confirmed_by == 'besart'
    assert row.confidence is None
    assert ImportMapping.query.count() == 1


def test_save_confirmed_mapping_stores_ai_confirmed_provenance(ctx, tmp_path):
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)
    wb = openpyxl.load_workbook(path)
    _set(wb['KiKxxl'], 9, 12, 88.0)
    wb.save(path)

    row = save_confirmed_mapping(
        file_type='tfc_forecast', filepath=path, mapping={'gesamt_col': 12},
        parse_fn=parse_tfc_file, resolution_source='ai_confirmed', confirmed_by='besart',
        confidence=0.87, rationale='label moved right',
    )

    assert row.resolution_source == 'ai_confirmed'
    assert row.confidence == 0.87
    assert row.rationale == 'label moved right'


def test_save_confirmed_mapping_upserts_same_file_type_and_fingerprint(ctx, tmp_path):
    path = _save(_tfc_workbook_without_gesamt(), tmp_path)
    wb = openpyxl.load_workbook(path)
    _set(wb['KiKxxl'], 9, 12, 88.0)
    wb.save(path)

    first = save_confirmed_mapping(
        file_type='tfc_forecast', filepath=path, mapping={'gesamt_col': 12},
        parse_fn=parse_tfc_file, resolution_source='manual', confirmed_by='besart',
    )
    second = save_confirmed_mapping(
        file_type='tfc_forecast', filepath=path, mapping={'gesamt_col': 12},
        parse_fn=parse_tfc_file, resolution_source='manual', confirmed_by='someone-else',
    )

    assert first.id == second.id
    assert ImportMapping.query.count() == 1
    assert ImportMapping.query.first().confirmed_by == 'someone-else'


def test_save_confirmed_mapping_rejects_a_mapping_that_still_warns(ctx, tmp_path):
    # No Monat/date columns at all — parse_tfc_file warns "No daily rows could
    # be read" regardless of what gesamt_col is given, since row detection
    # itself (not the mapping) is what fails here.
    wb = openpyxl.Workbook()
    wb.active.title = 'KiKxxl'
    path = _save(wb, tmp_path)

    try:
        save_confirmed_mapping(
            file_type='tfc_forecast', filepath=path, mapping={'gesamt_col': 30},
            parse_fn=parse_tfc_file, resolution_source='manual', confirmed_by='besart',
        )
        assert False, 'expected InvalidMapping'
    except InvalidMapping:
        pass

    assert ImportMapping.query.count() == 0
