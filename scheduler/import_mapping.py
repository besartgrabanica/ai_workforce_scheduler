"""
The 3-tier import resolution (docs/specs/2026-08-import-mapping-detection.md):
Tier 1 (the existing deterministic parser in scheduler/parsers/<project>.py,
untouched), Tier 2 (a previously confirmed ImportMapping applied directly —
resolve_cached_mapping/resolve_import below), and Tier 3 (an AI-proposed
mapping via scheduler.ai_import_mapping, never auto-applied). This module
also owns save_confirmed_mapping(), the only way an ImportMapping row gets
created or updated — used by app.py's admin-gated /import-mappings route to
confirm an AI proposal (edited or as-is) or to save a fully manual mapping.
"""
import hashlib
import json
from datetime import date, datetime

import openpyxl

from scheduler.ai_import_mapping import propose_mapping
from scheduler.models import ImportMapping, db


def compute_layout_fingerprint(filepath):
    """A stable hash of a workbook's structural shape: sheet names plus every
    *string*-valued cell in each sheet's first 8 rows. Both file types this
    feature covers keep their structural header labels within that range and
    their real per-period data below it or in numeric/date cells within it —
    e.g. eon.py's TFC parser only ever scans row 7 for the 'Gesamt' label,
    with actual daily figures starting row 9; its Abnahmemenge parser scans
    for a 'Datum' text header, with dates/numbers in the data cells beside
    it. Restricting to strings in the top rows means a new period's numbers
    (and even a first data row's own text, like a month/weekday label) never
    perturb the fingerprint — only the labels/positions that define the
    layout do. A client moving/renaming/reordering those labels produces a
    different fingerprint, which is exactly the invalidation this feature
    needs — see ImportMapping's (file_type, layout_fingerprint) uniqueness."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    parts = []
    for name in wb.sheetnames:
        parts.append(f'SHEET:{name}')
        ws = wb[name]
        for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
            for v in row:
                if isinstance(v, str) and v.strip():
                    parts.append(v.strip())
    return hashlib.sha256('\x1f'.join(parts).encode('utf-8')).hexdigest()


def resolve_cached_mapping(file_type, filepath):
    """Tier 2 lookup. Project scoping is implicit — ImportMapping lives in
    the active project-scoped DB, so this can only ever match a mapping
    confirmed under the currently active project. Returns the ImportMapping
    row, or None on a cache miss (no mapping yet, or the fingerprint no
    longer matches a reformatted file)."""
    fingerprint = compute_layout_fingerprint(filepath)
    return ImportMapping.query.filter_by(
        file_type=file_type, layout_fingerprint=fingerprint,
    ).first()


def _sheet_for_file_type(wb, file_type):
    """Same sheet each parser itself would pick — see parse_tfc_file (always
    the 'KiKxxl' sheet) and parse_abnahme_de_file (wb.active) in
    scheduler/parsers/eon.py."""
    if file_type == 'tfc_forecast' and 'KiKxxl' in wb.sheetnames:
        return wb['KiKxxl']
    return wb.active


def build_layout_snapshot(filepath, file_type, max_rows=15):
    """Sparse {'row': int, 'col': int, 'value': ...} cells (non-empty,
    JSON-serializable — dates/datetimes stringified) from the relevant
    sheet's first max_rows rows. This is Tier 3's raw material (see
    scheduler.ai_import_mapping.propose_mapping): wider than
    compute_layout_fingerprint's 8-row/string-only scope because the model
    also needs to *see* where numeric/date data actually starts to place a
    column with any confidence, not just read header labels."""
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = _sheet_for_file_type(wb, file_type)
    snapshot = []
    for r, row in enumerate(ws.iter_rows(min_row=1, max_row=max_rows, values_only=True)):
        for c, v in enumerate(row):
            if v is None:
                continue
            if isinstance(v, (datetime, date)):
                v = v.isoformat()
            snapshot.append({'row': r, 'col': c, 'value': v})
    return snapshot


def resolve_import(file_type, filepath, parse_fn):
    """Runs Tier 1 (parse_fn, unchanged — acceptance criterion 1) and, only
    if it reports a warning (a Tier 1 miss), looks up Tier 2. On a Tier 2
    hit that resolves cleanly, re-parses with the confirmed mapping applied
    and trusts that result instead. A Tier 2 hit that still produces
    warnings (the cached mapping no longer resolves this file cleanly
    either) is not trusted for parsing, and — like a plain Tier 2 miss —
    falls through to Tier 3.

    Tier 3 (scheduler.ai_import_mapping.propose_mapping) proposes a mapping
    but it is NEVER applied here: the returned result/warnings stay Tier
    1's own (degraded) output, and 'ai_proposed' + the proposal are
    returned alongside so a human can review it — see acceptance criterion
    2 ("nothing downstream executes pre-confirm"), enforced by the caller
    (app.py's forecast_new) refusing to proceed on this tier.

    parse_fn is a project parser's parse_tfc_file / parse_abnahme_de_file —
    both accept an optional `mapping=` keyword carrying a previously
    confirmed ImportMapping.mapping_data (JSON-decoded) that overrides the
    deterministic header/position search.

    Returns (result, warnings, tier, proposal) where tier is
    'deterministic', 'cached', or 'ai_proposed' (for ImportLog to record —
    acceptance criterion 6), and proposal is the dict returned by
    propose_mapping (or None unless tier == 'ai_proposed')."""
    result, warnings = parse_fn(filepath)
    if not warnings:
        return result, warnings, 'deterministic', None

    cached = resolve_cached_mapping(file_type, filepath)
    if cached is not None:
        cached_result, cached_warnings = parse_fn(filepath, mapping=json.loads(cached.mapping_data))
        if not cached_warnings:
            return cached_result, [], 'cached', None

    snapshot = build_layout_snapshot(filepath, file_type)
    proposal = propose_mapping(file_type, snapshot)
    if proposal is None:
        return result, warnings, 'deterministic', None
    return result, warnings, 'ai_proposed', proposal


class InvalidMapping(ValueError):
    """Raised by save_confirmed_mapping when the submitted mapping doesn't
    resolve filepath cleanly — the message is user-facing (flashed as-is)."""


def save_confirmed_mapping(file_type, filepath, mapping, parse_fn,
                            resolution_source, confirmed_by,
                            confidence=None, rationale=None):
    """The only way an ImportMapping row is created or updated. Never trusts
    the caller: re-parses filepath with `mapping` applied and only persists
    anything if that produces zero warnings (mirrors the same distrust
    resolve_import applies to a stale Tier 2 hit) — an admin can still
    submit a wrong mapping, but this refuses to silently save one that
    provably doesn't work against the very file they uploaded to prove it.

    resolution_source is 'manual' or 'ai_confirmed' (confirming — with or
    without edits — a Tier 3 proposal); confirmed_by is the admin's
    username (no FK — see ImportMapping's own docstring on why project-
    scoped models never reference the identity DB directly).

    Upserts keyed on (file_type, this file's layout_fingerprint) — a second
    confirm for the exact same layout updates the existing row rather than
    creating a duplicate; a different fingerprint (reformatted again, or a
    genuinely different file) always creates its own row.

    Raises InvalidMapping (message is safe to flash to the user) instead of
    saving when the mapping doesn't resolve cleanly."""
    _, warnings = parse_fn(filepath, mapping=mapping)
    if warnings:
        raise InvalidMapping(
            "This mapping doesn't produce a clean parse of the uploaded file: "
            + ' '.join(warnings)
        )

    fingerprint = compute_layout_fingerprint(filepath)
    row = ImportMapping.query.filter_by(file_type=file_type, layout_fingerprint=fingerprint).first()
    if row is None:
        row = ImportMapping(file_type=file_type, layout_fingerprint=fingerprint)
        db.session.add(row)

    row.mapping_data = json.dumps(mapping)
    row.resolution_source = resolution_source
    row.confidence = confidence
    row.rationale = rationale
    row.confirmed_by = confirmed_by
    db.session.commit()
    return row
