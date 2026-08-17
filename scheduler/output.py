"""
Export a Schedule to an Excel workbook with two sheets:
  1. Schedule  – employee × day calendar grid (colour-coded)
  2. Coverage  – daily required vs. scheduled agents + stats
"""
import calendar
from datetime import date
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from .models import Employee, Schedule, ShiftAssignment

_MONTH_NAMES = [
    '', 'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December',
]

_THIN = Side(style='thin', color='CCCCCC')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)

_STATUS_FILLS = {
    'day_off':      PatternFill('solid', fgColor='CBD5E1'),
    'weekend_off':  PatternFill('solid', fgColor='E2E8F0'),
    'holiday_off':  PatternFill('solid', fgColor='FCA5A5'),
    'constraint_off': PatternFill('solid', fgColor='F1F5F9'),
}

_HEADER_FILL  = PatternFill('solid', fgColor='1E3A5F')
_WEEKEND_FILL = PatternFill('solid', fgColor='E2E8F0')
_HOLIDAY_FILL = PatternFill('solid', fgColor='FEE2E2')
_WHITE_FILL   = PatternFill('solid', fgColor='FFFFFF')

_HEADER_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
_LABEL_FONT   = Font(bold=True, name='Calibri', size=9)
_CELL_FONT    = Font(name='Calibri', size=8)
_SMALL_FONT   = Font(name='Calibri', size=7)


def _hex_fill(hex_color: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_color.lstrip('#'))


def _cell(ws, row, col, value='', font=None, fill=None, align='center', border=True, wrap=False):
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if border:
        c.border = _BORDER
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    return c


def export_schedule(schedule_id: int, project_display_name: str = 'Workforce') -> BytesIO:
    sched = Schedule.query.get(schedule_id)
    year, month = sched.year, sched.month
    _, days_in_month = calendar.monthrange(year, month)
    all_days = [date(year, month, d) for d in range(1, days_in_month + 1)]

    # Gather all assignments indexed by (employee_id, date)
    assignments: dict[tuple, ShiftAssignment] = {}
    for a in ShiftAssignment.query.filter_by(schedule_id=schedule_id).all():
        assignments[(a.employee_id, a.date)] = a

    employees = (Employee.query
                 .filter_by(status='active')
                 .order_by(Employee.team, Employee.name)
                 .all())

    wb = openpyxl.Workbook()

    # ── Sheet 1: Schedule grid ──────────────────────────────────────────────
    ws = wb.active
    ws.title = 'Schedule'
    ws.sheet_view.showGridLines = False

    # Title
    ws.merge_cells(f'A1:{get_column_letter(days_in_month + 3)}1')
    t = ws['A1']
    t.value = f'{project_display_name} Workforce Schedule — {_MONTH_NAMES[month]} {year}'
    t.font = Font(bold=True, name='Calibri', size=13, color='1E3A5F')
    t.alignment = Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 22

    # Header row: Team | Name | FTE | day-1 … day-N
    ws.row_dimensions[2].height = 28
    _cell(ws, 2, 1, 'Team',  font=_HEADER_FONT, fill=_HEADER_FILL)
    _cell(ws, 2, 2, 'Name',  font=_HEADER_FONT, fill=_HEADER_FILL, align='left')
    _cell(ws, 2, 3, 'FTE %', font=_HEADER_FONT, fill=_HEADER_FILL)

    day_headers = []
    for i, day in enumerate(all_days, start=4):
        dow = day.weekday()
        day_names = ['Mo', 'Tu', 'We', 'Th', 'Fr', 'Sa', 'Su']
        label = f'{day_names[dow]}\n{day.day}'
        is_we = dow >= 5
        fill = _WEEKEND_FILL if is_we else _HEADER_FILL
        font = Font(bold=True, name='Calibri', size=8, color='475569' if is_we else 'FFFFFF')
        c = _cell(ws, 2, i, label, font=font, fill=fill)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        day_headers.append((i, day))

    # Column widths
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 22
    ws.column_dimensions['C'].width = 6
    for col_i, _ in day_headers:
        ws.column_dimensions[get_column_letter(col_i)].width = 8

    # Data rows
    prev_team = None
    row = 3
    for emp in employees:
        if emp.team != prev_team:
            # Team separator row
            ws.merge_cells(f'A{row}:{get_column_letter(days_in_month + 3)}{row}')
            tc = ws[f'A{row}']
            tc.value = emp.team or '—'
            tc.font = Font(bold=True, name='Calibri', size=9, color='FFFFFF')
            tc.fill = PatternFill('solid', fgColor='334155')
            tc.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[row].height = 14
            row += 1
            prev_team = emp.team

        ws.row_dimensions[row].height = 16
        _cell(ws, row, 1, emp.team or '', font=_CELL_FONT, fill=_WHITE_FILL, align='left')
        _cell(ws, row, 2, emp.name,       font=_CELL_FONT, fill=_WHITE_FILL, align='left')
        _cell(ws, row, 3, f'{emp.fte_percent}%', font=_SMALL_FONT, fill=_WHITE_FILL)

        for col_i, day in day_headers:
            a = assignments.get((emp.id, day))
            if a is None:
                _cell(ws, row, col_i, '', fill=_WHITE_FILL)
                continue

            text = a.display_code
            if a.status == 'work' and a.shift_template and a.shift_template.color:
                fill = _hex_fill(a.shift_template.color)
                font = Font(name='Calibri', size=7, bold=True, color='FFFFFF')
            elif a.status in _STATUS_FILLS:
                fill = _STATUS_FILLS[a.status]
                font = _SMALL_FONT
            else:
                fill = _WHITE_FILL
                font = _SMALL_FONT

            c = _cell(ws, row, col_i, text, font=font, fill=fill)
            c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        row += 1

    # Freeze panes at D3
    ws.freeze_panes = 'D3'

    # ── Sheet 2: Coverage ───────────────────────────────────────────────────
    ws2 = wb.create_sheet('Coverage')
    ws2.sheet_view.showGridLines = False

    hdr = ['Date', 'Day', 'Required Agents', 'Scheduled Agents', 'Coverage %', 'Gap']
    ws2.row_dimensions[1].height = 20
    for ci, h in enumerate(hdr, 1):
        _cell(ws2, 1, ci, h, font=_HEADER_FONT, fill=_HEADER_FILL)

    ws2.column_dimensions['A'].width = 12
    ws2.column_dimensions['B'].width = 8
    ws2.column_dimensions['C'].width = 16
    ws2.column_dimensions['D'].width = 16
    ws2.column_dimensions['E'].width = 12
    ws2.column_dimensions['F'].width = 8

    day_names_long = ['Mon', 'Tue', 'Wed', 'Thu', 'Fri', 'Sat', 'Sun']

    for r, day in enumerate(all_days, start=2):
        # Count scheduled workers for this day
        day_assignments = [
            a for a in assignments.values()
            if a.date == day and a.status == 'work'
        ]
        scheduled = len(day_assignments)

        # Required from DB or default
        from .models import DailyForecast
        fc = DailyForecast.query.filter_by(date=day).first()
        required = (fc.required_ks_agents if fc and fc.required_ks_agents else
                    (220 if day.weekday() < 5 else 55 if day.weekday() == 5 else 0))

        pct = round(scheduled / required * 100, 1) if required else 100.0
        gap = scheduled - required

        is_we = day.weekday() >= 5
        row_fill = _WEEKEND_FILL if is_we else _WHITE_FILL

        _cell(ws2, r, 1, day.strftime('%d.%m.%Y'), font=_CELL_FONT, fill=row_fill)
        _cell(ws2, r, 2, day_names_long[day.weekday()], font=_CELL_FONT, fill=row_fill)
        _cell(ws2, r, 3, required,  font=_CELL_FONT, fill=row_fill)
        _cell(ws2, r, 4, scheduled, font=_CELL_FONT, fill=row_fill)

        pct_fill = (PatternFill('solid', fgColor='BBF7D0') if pct >= 100
                    else PatternFill('solid', fgColor='FEF08A') if pct >= 80
                    else PatternFill('solid', fgColor='FCA5A5'))
        _cell(ws2, r, 5, f'{pct}%', font=_CELL_FONT, fill=pct_fill)

        gap_fill = (PatternFill('solid', fgColor='BBF7D0') if gap >= 0
                    else PatternFill('solid', fgColor='FCA5A5'))
        _cell(ws2, r, 6, gap, font=_CELL_FONT, fill=gap_fill)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
