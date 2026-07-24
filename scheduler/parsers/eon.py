"""
EON's forecast/employee-file parsers — registered for project key 'eon' in
scheduler/parsers/__init__.py. Parses the three source Excel files:
  1. TFC_Juni-August 2026_KiKxxl.xlsx  – half-hourly E.ON contact forecast
  2. Abnahmemenge DE Juni.xlsx         – German offices daily contribution
  3. Specifikat e punetoreve - E.ON.xlsx – Kosovo employee list
plus the Forecast Calculation file (Kosovo required-agent counts) below.

These four functions and their docstrings are the interface contract for any
other project's parser module: onboarding a new client means writing a new
module in this package that implements whichever of these apply to that
client's real files (there's no way to auto-detect an arbitrary external
Excel layout — see scheduler/parsers/__init__.py and RUNBOOK.md for the
process), not touching this file.
"""
import openpyxl
from datetime import datetime, date


# ── TFC forecast ────────────────────────────────────────────────────────────

# The "Gesamt" (total) Forecast section starts at column 337 (1-based).
# Layout within that section (1-based column offsets from col 337):
#   +0  Synchron daily total
#   +1 … +24  30-min slots 08:00–19:30
#   +25 Asynchron daily total
#   +26 … +49  30-min slots
#   +50 Chat daily total
#   +51 … +74  30-min slots
#   +75 Grand total (Gesamt)
_GESAMT_SYNC_OFFSET = 0    # relative to section start (0-indexed)
_GESAMT_ASYNC_OFFSET = 25
_GESAMT_CHAT_OFFSET = 50
_SLOT_COUNT = 24           # 08:00 … 19:30 in 30-min steps

_TIME_SLOTS = [
    f'{8 + i // 2:02d}:{(i % 2) * 30:02d}' for i in range(_SLOT_COUNT)
]


def _find_section_col(ws, section_name):
    """Find 0-based column index for a section label in row 7."""
    row7 = next(ws.iter_rows(min_row=7, max_row=7, values_only=True))
    for i, v in enumerate(row7):
        if str(v).strip() == section_name:
            return i
    return None


def _parse_date_cell(val):
    if val is None:
        return None
    if isinstance(val, (datetime,)):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        for fmt in ('%d.%m.%Y', '%Y-%m-%d'):
            try:
                return datetime.strptime(val.strip(), fmt).date()
            except ValueError:
                continue
    return None


def parse_tfc_file(filepath):
    """
    Returns (results, warnings).
    results is a list of dicts, one per calendar day:
      {
        'date': date,
        'day_tag': str,      e.g. 'Mo'
        'kw': int,
        'total_sync': float,
        'total_async': float,
        'total_chat': float,
        'slots': {'HH:MM': {'sync': float, 'async': float, 'chat': float}, ...}
      }
    warnings is a list of human-readable strings describing any expected
    structure (sheet name, header row) that wasn't found — signals the source
    file's format may have changed since this parser was written.
    """
    warnings = []
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    if 'KiKxxl' not in wb.sheetnames:
        raise ValueError(
            f"Expected a sheet named 'KiKxxl' but found: {', '.join(wb.sheetnames)}. "
            "The file format may have changed."
        )
    ws = wb['KiKxxl']

    gesamt_col = _find_section_col(ws, 'Gesamt')
    if gesamt_col is None:
        # Fallback: hard-coded 0-indexed position from our file analysis
        gesamt_col = 336
        warnings.append(
            "Expected a 'Gesamt' (total) header in row 7 was not found — fell back to a "
            "hardcoded column position. Please double-check the imported totals are correct; "
            "the file's column layout may have changed."
        )

    results = []
    for row in ws.iter_rows(min_row=9, values_only=True):
        monat = row[0]
        if not monat:
            continue

        d = _parse_date_cell(row[3])
        if d is None:
            continue

        def _get(offset):
            idx = gesamt_col + offset
            return float(row[idx] or 0) if idx < len(row) else 0.0

        total_sync = _get(_GESAMT_SYNC_OFFSET)
        total_async = _get(_GESAMT_ASYNC_OFFSET)
        total_chat = _get(_GESAMT_CHAT_OFFSET)

        slots = {}
        for i, t in enumerate(_TIME_SLOTS):
            slots[t] = {
                'sync':  _get(_GESAMT_SYNC_OFFSET + 1 + i),
                'async': _get(_GESAMT_ASYNC_OFFSET + 1 + i),
                'chat':  _get(_GESAMT_CHAT_OFFSET + 1 + i),
            }

        results.append({
            'date': d,
            'day_tag': str(row[2] or ''),
            'kw': int(row[1]) if row[1] else None,
            'total_sync': total_sync,
            'total_async': total_async,
            'total_chat': total_chat,
            'slots': slots,
        })

    if not results:
        warnings.append(
            "No daily rows could be read starting at row 9 — the date column may have moved. "
            "The file's row/column layout may have changed."
        )

    return results, warnings


# ── Abnahmemenge DE (German offices) ────────────────────────────────────────

def parse_abnahme_de_file(filepath):
    """
    Returns (result, warnings). result is a dict date → germany_daily_contribution
    (Abnahme DE netto). warnings flags if the expected 'Datum' header wasn't found,
    which means the file's format may have changed.
    """
    warnings = []
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    # Find header row (contains 'Datum')
    header_row = None
    abnahme_de_col = None
    for r in ws.iter_rows(values_only=True):
        for i, v in enumerate(r):
            if str(v).strip() == 'Datum':
                header_row = r
                # 'Abnahme DE netto' is the last non-None column in header
                for j in range(len(r) - 1, i, -1):
                    if r[j] is not None:
                        abnahme_de_col = j
                        break
                break
        if header_row is not None:
            break

    if abnahme_de_col is None:
        warnings.append(
            "Expected a 'Datum' header was not found in this file — no daily contribution "
            "data was extracted. The file's format may have changed."
        )
        return {}, warnings

    result = {}
    for row in ws.iter_rows(values_only=True):
        d = _parse_date_cell(row[1]) if len(row) > 1 else None
        if d is None:
            continue
        val = row[abnahme_de_col] if abnahme_de_col < len(row) else None
        if val is not None:
            try:
                result[d] = float(val)
            except (TypeError, ValueError):
                pass

    return result, warnings


# ── Employee specification ────────────────────────────────────────────────────

def parse_employee_file(filepath):
    """
    Returns (employees, warnings). employees is a list of raw employee dicts:
      {
        'name': str,
        'team': str,
        'raw_notes': str,
        'works_saturday': str,   # 'yes'/'no'/''
        'works_sunday': str,     # 'yes'/'no'/''
        'works_holidays': bool,
        'fte_fraction': float,   # 0.0–1.0; None if not specified
      }
    warnings flags if the expected 'Emri & Mbiemri' header row was never seen, or
    no employees were found at all — both signal the column layout may have changed.
    """
    warnings = []
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    current_team = ''
    employees = []
    saw_expected_header = False

    for row in ws.iter_rows(values_only=True):
        team_cell = row[0]
        name_cell = row[1]
        notes_cell = row[2]
        weekend_cell = row[3]
        holiday_cell = row[4]
        contract_cell = row[5]

        if team_cell and str(team_cell).strip().startswith('Team'):
            current_team = str(team_cell).strip()

        if name_cell and str(name_cell).strip() == 'Emri & Mbiemri':
            saw_expected_header = True

        if not name_cell or str(name_cell).strip() in ('Emri & Mbiemri', 'Note:', ''):
            continue

        name = str(name_cell).strip()
        if not name or name.lower() in ('note:',):
            continue

        # Weekend flags — split into independent Saturday/Sunday availability.
        # Historical 'YES' predates Sunday being a workable day at all, so it's
        # read as Saturday-yes (Sunday-no); explicit 'Saturday only'/'Sunday only'
        # map to exactly one of the two days.
        we_raw = str(weekend_cell).strip().lower() if weekend_cell else ''
        if we_raw == 'yes':
            works_saturday, works_sunday = 'yes', 'no'
        elif we_raw == 'no':
            works_saturday, works_sunday = 'no', 'no'
        elif we_raw == 'saturday only':
            works_saturday, works_sunday = 'yes', 'no'
        elif we_raw == 'sunday only':
            works_saturday, works_sunday = 'no', 'yes'
        elif we_raw == '':
            works_saturday, works_sunday = '', ''
        else:
            works_saturday, works_sunday = 'yes', 'no'

        # Holiday flag
        hol_raw = str(holiday_cell).strip().lower() if holiday_cell else ''
        works_holidays = hol_raw == 'yes'

        # FTE
        fte_fraction = None
        if contract_cell is not None:
            try:
                fte_fraction = float(contract_cell)
            except (TypeError, ValueError):
                pass

        employees.append({
            'name': name,
            'team': current_team,
            'raw_notes': str(notes_cell).strip() if notes_cell else '',
            'works_saturday': works_saturday,
            'works_sunday': works_sunday,
            'works_holidays': works_holidays,
            'fte_fraction': fte_fraction,
        })

    if not saw_expected_header:
        warnings.append(
            "Expected header 'Emri & Mbiemri' was not found in the name column — confirm "
            "column order still matches Team | Name | Notes | Weekend | Holiday | Contract. "
            "The file's format may have changed."
        )
    if not employees:
        warnings.append("No employees were found in this file.")

    return employees, warnings


# ── Forecast Calculation (Kosovo FTE, optional) ─────────────────────────────

def parse_forecast_calculation_file(filepath):
    """
    Returns (result, warnings). result is a dict date → required_ks_agents
    (integer FTE count for Kosovo), reading the 'KiKxxl\nPrishtina' column.
    warnings flags if that column wasn't found, meaning the file's format may
    have changed.
    """
    warnings = []
    wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
    ws = wb.active

    # Find 'KiKxxl\nPrishtina' or 'KiKxxl' in the header row
    ks_col = None
    date_col = None
    for r in ws.iter_rows(values_only=True):
        for i, v in enumerate(r):
            if v is not None and 'kikxxl' in str(v).lower():
                ks_col = i
            if v is not None and str(v).strip() == 'Datum':
                date_col = i
        if ks_col is not None:
            break

    if ks_col is None:
        warnings.append(
            "Expected a column containing 'KiKxxl' was not found — no Kosovo agent "
            "requirements were extracted from this file. The file's format may have changed."
        )
        return {}, warnings

    result = {}
    for row in ws.iter_rows(min_row=10, values_only=True):
        if not row or len(row) <= max(ks_col, date_col or 1):
            continue
        d = _parse_date_cell(row[date_col]) if date_col is not None else _parse_date_cell(row[1])
        if d is None:
            continue
        val = row[ks_col]
        if val is not None:
            try:
                result[d] = int(float(val))
            except (TypeError, ValueError):
                pass

    return result, warnings
